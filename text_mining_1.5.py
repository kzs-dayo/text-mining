import tkinter as tk
import tkinter.ttk as ttk
import tkinter.filedialog as filedialog
import tkinter.messagebox as tkm
import xlrd
import openpyxl
import re
import unicodedata
import pickle
import os
import gc
import csv
import urllib
import mojimoji
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import collections
from collections import Counter
from gensim.models.phrases import Phrases , Phraser
from janome.tokenizer import Tokenizer
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.feature_extraction.text import TfidfTransformer
from sklearn.decomposition import PCA
from sklearn.cluster import KMeans
from functools import partial
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Alignment, Font

class Clustering():
    """
    クラスタリングに関する処理を行う
    """
    def read_file(self):
        """
        指定のファイルを読み込んで、列の値を返す

        """
        global data_file_path
        global save_file
        global status
        global column
        global sheet_name
        global arr_data
        global arr_origin_data

        book = xlrd.open_workbook(data_file_path)
        sheet = book.sheet_by_name(sheet_name)

        data_list = []
        origin_data = []
        #1行ずつ処理していく
        for row in range(sheet.nrows):
            cell_value = sheet.cell(row, column).value
            origin_data.append(cell_value)
            #アルファベットの大文字を小文字に変換
            cell_value = cell_value.lower()
            #全角と半角を変換
            cell_value = mojimoji.zen_to_han(cell_value, kana = False)
            cell_value = mojimoji.han_to_zen(cell_value, ascii=False, digit=False)
            #スペースを削除
            cell_value = re.sub(r'[ 　]', "", cell_value)
            #濁点や半濁点が分離しているのを直前の文字と結合する
            cell_value = self.join_diacritic(cell_value)
            #記号になっている超音符を修正する
            cell_value = self.choonpu_unity(cell_value)

            data_list.append(cell_value)
        #カラム行を削除
        del data_list[0]
        del origin_data[0]
        #arrayに変換(メモリ節約のため)
        arr_data = np.array(data_list)
        arr_origin_data = np.array(origin_data)

        if status == "追加":
            exist_origin_data = np.load(save_file + '/origin_data.npy')
            arr_origin_data = np.concatenate([arr_origin_data, exist_origin_data])

        np.save(save_file + '/origin_data.npy', arr_origin_data)

        #メモリ解放
        del data_list
        del origin_data
        gc.collect()

    def join_diacritic(self, text, mode = "NFC"):
        """
        濁点や半濁点が分離しているのを直前の文字と結合する
        """
        # str -> bytes
        bytes_text = text.encode()

        # 濁点Unicode結合文字置換
        bytes_text = re.sub(b"\xe3\x82\x9b", b'\xe3\x82\x99', bytes_text)
        bytes_text = re.sub(b"\xef\xbe\x9e", b'\xe3\x82\x99', bytes_text)

        # 半濁点Unicode結合文字置換
        bytes_text = re.sub(b"\xe3\x82\x9c", b'\xe3\x82\x9a', bytes_text)
        bytes_text = re.sub(b"\xef\xbe\x9f", b'\xe3\x82\x9a', bytes_text)

        # bytet -> str
        text = bytes_text.decode()

        # 正規化
        text = unicodedata.normalize(mode, text)

        return text

    def choonpu_unity(self, text):
        """
        記号になっている超音符（例：ラーメンの伸ばし棒）を置換する
        """
        text = list(text)
        for n in range(len(text)-1):
            kana = re.match('[ァ-ン]', text[n])
            if kana:
              haihun = re.match('-', text[n+1])
              if haihun:
                text[n+1] = 'ー'
        text = "".join(text)

        return text
        
    def morphological(self):
        """
        形態素解析を行う
        """
        global arr_data
        global all_token
        global userdict_file_path
        all_token = []

        #SlothLibのテキストページ を取得して、ストップワード辞書を作成
        url = 'http://svn.sourceforge.jp/svnroot/slothlib/CSharp/Version1/SlothLib/NLP/Filter/StopWord/word/Japanese.txt'
        slothlib_file = urllib.request.urlopen(url=url)
        slothlib_stopwords = [line.decode("utf-8").strip() for line in slothlib_file]
        slothlib_stopwords = [ss for ss in slothlib_stopwords if not ss==u'']

        if userdict_file_path:
            #ユーザー辞書を使用
            t = Tokenizer(userdict_file_path, udic_enc='cp932')
        else:
            #ユーザー辞書非使用
            t = Tokenizer()

        for data in arr_data.tolist():
            each_token_dict = {}
            #形態素解析を実行
            tokens = t.tokenize(data)
            for token in tokens:
                #スペースを削除
                token.base_form = re.sub(r'[ 　]', "", token.base_form)
                if token.base_form != "" and token.base_form not in slothlib_stopwords:
                    #単語(基本形)と品詞の辞書を作成
                    each_token_dict[token.base_form] = token.part_of_speech.split(',')[0]
            all_token.append(each_token_dict)

    def digitization(self):
        """
        ベクトル化と重み付けを行う
        Returns
        -------
        pca_data : 
            ベクトル化と重み付けを行ったデータ
        """
        global segmentation_datas
        #ベクトル化を実行
        count = CountVectorizer()
        bags = count.fit_transform(segmentation_datas)

        #重み付けを実行
        tfidf = TfidfTransformer()
        np.set_printoptions(precision = 4)
        tf_idf = tfidf.fit_transform(bags)

        #次元の削減
        pca = PCA(n_components = 2)
        pca_data = pca.fit_transform(tf_idf.toarray())

        print(pca.explained_variance_ratio_)
        print(pca.components_)
        print(pca_data)

        return pca_data

    def clustering(self, pca_data, k_value):
        """
        K-meansクラスタリングを行う

        Parameters
        ----------
        pca_data : 
            ベクトル化と重み付けを行ったデータ
        kvalue：int
            K値

        Returns
        -------
        labels : ndarray
            クラスタ番号
        sse : 
            SSE値
        ranking_tops : list
            上位１０個のクラスタ番号
        """
        #指定されたクラスタ数で、クラスタリングを実行
        #k-Means++法によりクラスタ中心点を選択
        km = KMeans(n_clusters = k_value, init = 'k-means++', random_state = 0).fit(pca_data)
        
        #クラスタの中心点
        centers = km.cluster_centers_
        #SSE値
        sse = km.inertia_

        labels = km.predict(pca_data)

        #クラスタ情報をファイルに出力する
        with open(save_file + '/sample.pickle', mode="wb") as f:
            pickle.dump(km, f)
        np.save(save_file + '/label_data.npy', labels)
        np.save(save_file + '/pca_data.npy', pca_data)

        #クラスタリング結果のプロット
        plt.scatter(pca_data[:,0], pca_data[:,1], c = labels)
        plt.scatter(centers[:,0], centers[:,1], marker = '*', c = 'red')
        plt.show(block=False)

        #データの多い順トップ１０のクラスタを抽出していく
        count_dict = Counter(labels.tolist()) #各クラスタの出現頻度を数える
        ranking_list = sorted(count_dict.items(), key = lambda x:x[1], reverse=True) #多い順でソート
        ranking_tops = []
        if len(ranking_list) >= 10:
            for i in range(10): #出現頻度がトップ10のクラスタだけ抽出
                ranking = ranking_list[i]
                ranking_tops.append(ranking[0])
        else:
            for i in range(len(ranking_list)):
                ranking = ranking_list[i]
                ranking_tops.append(ranking[0])

        return labels, sse, ranking_tops

    def assign_cluster(self, pca_data):
        """
        既存のクラスタにデータを割り当てる
        """
        #クラスタ情報をファイルから読み込む
        with open(save_file + '/sample.pickle', mode="rb") as f:
            km = pickle.load(f)
        exist_label = np.load(save_file + '/label_data.npy')
        exist_pca = np.load(save_file + '/pca_data.npy')

        #クラスタの中心点
        centers = km.cluster_centers_
        #SSE値
        sse = km.inertia_

        new_label = km.predict(pca_data)

        labels = np.concatenate([exist_label, new_label])
        pca_data = np.concatenate([exist_pca, pca_data])

        np.save(save_file + '/label_data.npy', labels)
        np.save(save_file + '/pca_data.npy', pca_data)

        #クラスタリング結果のプロット
        plt.scatter(pca_data[:,0], pca_data[:,1], c = labels)
        plt.scatter(centers[:,0], centers[:,1], marker = '*', c = 'red')
        plt.show(block=False)

        #データの多い順トップ１０のクラスタを抽出していく
        count_dict = Counter(labels.tolist()) #各クラスタの出現頻度を数える
        ranking_list = sorted(count_dict.items(), key = lambda x:x[1], reverse=True) #多い順でソート
        ranking_tops = []
        if len(ranking_list) >= 10:
            for i in range(10): #出現頻度がトップ10のクラスタだけ抽出
                ranking = ranking_list[i]
                ranking_tops.append(ranking[0])
        else:
            for i in range(len(ranking_list)):
                ranking = ranking_list[i]
                ranking_tops.append(ranking[0])

        return labels, sse, ranking_tops


    def output_excel(self, labels, sse, ranking_tops):
        """
        クラスタリング結果と上位１０個のクラスタをExcelに出力する

        Parameters
        ----------
        labels : 
            クラスタ番号
        sse : 
            SSE値
        ranking_tops : list
            上位１０個のクラスタ番号
        """
        #クラスタ番号とデータを出力
        global arr_origin_data
        wb = openpyxl.Workbook()
        wb.worksheets[0].title = "result"
        bs = wb["result"]
        #セル設定
        alignment = Alignment(horizontal='general',
                    vertical='bottom',
                    text_rotation=0,
                    wrap_text=True,
                    shrink_to_fit=False,
                    indent=0)
        bs.cell(1, 1).value = "クラスタリング結果"
        #SSE値の出力
        bs.cell(2, 1).value = "SSE値：" + str(sse)
        bs.cell(3, 1).value = "クラスター"
        bs.cell(3, 2).value = "読込データ"
        for row, (cluster, data) in enumerate(zip(labels, arr_origin_data.tolist()), start = 4):
            bs.cell(row, 1).value = cluster
            bs.cell(row, 2).value = data
            bs.cell(row, 2).alignment = alignment
        #テーブルを挿入
        n = bs.max_row
        m = bs.max_column
        p = self.num2alpha(m)
        '%s%d' % (p,n)
        o = 'A3:%s%d' %(p,n)
        tab = Table(displayName="Table1", ref= o)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=False)
        tab.tableStyleInfo = style
        bs.add_table(tab)
        #セル幅の調整
        bs.column_dimensions['B'].width = 60
        #フォントの設定
        font = Font(name='游ゴシック')
        for row in bs:
            for cell in row:
                bs[cell.coordinate].font = font
        bs.cell(1, 1).font = openpyxl.styles.Font(size=18, bold=True)
        wb.save('./result.xlsx')

        #ランキングを出力
        wb = openpyxl.load_workbook('./result.xlsx')
        bs = wb.create_sheet("ranking")
        bs = wb["ranking"]
        bs.cell(1, 1).value = "データ数ランキング"
        for i in range(len(ranking_tops)):
            ranking_top = ranking_tops[i]
            bs.cell(i + 2, 1).value = str(i+1) + "位"
            bs.cell(i + 2, 2).value = "クラスタ" + str(ranking_top)
        #フォントの設定
        font = Font(name='游ゴシック')
        for row in bs:
            for cell in row:
                bs[cell.coordinate].font = font
        bs.cell(1, 1).font = openpyxl.styles.Font(size=18, bold=True)
        wb.save('./result.xlsx')

    def elbow(self, pca_data, elbow_kvalue):
        """
        エルボー法を実行する

        Parameters
        ---------
        pca_data : 
            ベクトル化と重み付けを行ったデータ
        elbow_kvalue：int
            エルボー法を実行する上での最大のK値
        """
        distortions = []
        for i in range(1, elbow_kvalue):
            km = KMeans(n_clusters = i, init = 'k-means++').fit(pca_data)
            distortions.append(km.inertia_)
        plt.plot(range(1, elbow_kvalue), distortions, marker = 'o')
        plt.show()

    def num2alpha(self, num):
        """
        数字をアルファベットに変換
        """
        if num<=26:
            return chr(64+num)
        elif num%26==0:
            return num2alpha(num//26-1)+chr(90)
        else:
            return num2alpha(num//26)+chr(64+num%26)

class EntryForm(ttk.Frame):
    """
    入力フォームを表示し、入力された値を受け取る
    """
    def __init__(self, root):
        super().__init__(root)

    def create_widget(self):
        """
        入力フォームのウィジェットを配置
        """
        self.top_frame = tk.Frame(root)
        self.top_frame.pack(fill = tk.BOTH)

        #画面名ラベル
        title_label = ttk.Label(self.top_frame, text = "初期設定", font = ('', 20))
        title_label.pack(padx = 5, pady = (30, 0))
        subtitle_label = ttk.Label(self.top_frame, text = "初期設定を行います。")
        subtitle_label.pack(pady = 5)

        main_frame = tk.Frame(self.top_frame)
        main_frame.pack(pady = 130)

        button_frame = tk.Frame(self.top_frame)
        button_frame.pack(anchor = tk.SE, padx =30)

        self.err_frame = tk.Frame(self.top_frame)
        self.err_frame.pack()

        label = ttk.Label(main_frame, text = "〇実行内容")
        label.grid(row = "0", column = "0", sticky = tk.W)
        self.combo = ttk.Combobox(main_frame, state = "readonly")
        self.combo["values"] = ("新規","追加")
        self.combo.current(0)
        self.combo.grid(row = "1", column = "0", sticky = tk.W)
        label1 = ttk.Label(main_frame, text = "新規：新規でクラスタを作成する\n追加：既存のクラスタにデータを割り当てる")
        label1.grid(row = "2", column = "0", columnspan = "3", sticky = tk.W, pady = (0, 20))

        #読込ファイルテキストボックス
        label2 = ttk.Label(main_frame, text = "〇読込データの指定\nファイルパス")
        label2.grid(row = "3", column = "0", columnspan = "2", sticky = tk.W)
        self.txt_file_path = ttk.Entry(main_frame, width = 50)
        self.txt_file_path.grid(row = "4", column = "0", columnspan = "3", sticky = tk.W)
        button_selectfile = ttk.Button(main_frame, text = "参照", command = self.select_file)
        button_selectfile.grid(row = "4", column = "3", sticky = tk.W)

        #読込シートテキストボックス
        label3 = ttk.Label(main_frame, text = "シート名")
        label3.grid(row = "5", column = "0", sticky = tk.W)
        self.txt_sheet_name = ttk.Entry(main_frame, width = 20)
        self.txt_sheet_name.grid(row = "5", column = "1", sticky = tk.W)

        #読込列テキストボックス
        label4 = ttk.Label(main_frame, text = "列(数字で入力)")
        label4.grid(row = "6", column = "0", sticky = tk.W, pady = (2, 20))
        self.txt_column = ttk.Entry(main_frame, width = 3)
        self.txt_column.grid(row = "6", column = "1", sticky = tk.W, pady = (2, 20))

        label5 = ttk.Label(main_frame, text = "〇セーブデータの保存場所")
        label5.grid(row = "7", column = "0", columnspan = "2", sticky = tk.W)
        self.txt_save_file = ttk.Entry(main_frame, width = 50)
        self.txt_save_file.grid(row = "8", column = "0", columnspan = "3", sticky = tk.W)

        button_selectdir = ttk.Button(main_frame, text = "参照", command = self.select_dir)
        button_selectdir.grid(row = "8", column = "3", sticky = tk.W)
        label5 = ttk.Label(main_frame, text = "クラスタの情報や使用したデータなどを保存します。\nデータを追加割当する際に使用します。")
        label5.grid(row = "9", column = "0", columnspan = "3", sticky = tk.W)

        #次へボタン
        button_morphological = ttk.Button(button_frame, text = "次へ", command = partial(self.call_morphological_func, root))
        button_morphological.pack()

        self.error_label = tk.Label(self.err_frame)
        self.error_label.pack()
    
    def select_file(self):
        """
        参照ボタン押下時の処理
        """
        iDir = os.path.abspath(os.path.dirname(__file__))
        file_path = tk.filedialog.askopenfilename(initialdir = iDir)
        self.txt_file_path.insert(tk.END, file_path)

    def select_dir(self):
        """
        参照ボタン押下時の処理
        """
        dir = os.path.abspath(os.path.dirname(__file__))
        file_path = tk.filedialog.askdirectory(initialdir = dir) 
        self.txt_save_file.insert(tk.END, file_path)

    def call_morphological_func(self, root):
        """
        次へボタン押下時の処理
        """
        self.err_message = []

        try:
            #入力値の取得
            self.get_input_value()
            #ファイルの読み込み
            clustering = Clustering()
            clustering.read_file()
        except ValueError as e:
            self.err_message.append("ERROR2:列は数字で入力してください。")
        except FileNotFoundError as e:
            self.err_message.append("ERROR3:ご指定のファイルが見つかりません。ファイルパスを確認してください。")
        except xlrd.biffh.XLRDError as e:
            self.err_message.append("ERROR4:ご指定のシートが見つかりません。シート名を確認してください。")
        except IndexError as e:
            self.err_message.append("ERROR13:ご指定の列にデータがありません。")
        except Exception as e:
            self.err_message.append(str(e))
        else:
            self.top_frame.destroy()
            ud = DataCleansing(root)
            ud.create_frame()

        if self.err_message:
            self.error_label.configure(text = "\n".join(self.err_message), fg = "red")

    def get_input_value(self):
        """
        GUIで入力された値を変数に格納する
        """
        global data_file_path
        global column
        global sheet_name
        global save_file
        global status

        data_file_path = self.txt_file_path.get() #読み込みファイルのパス
        column = self.txt_column.get() #読み込む列
        sheet_name = self.txt_sheet_name.get() #読み込むシート
        save_file = self.txt_save_file.get()
        status = self.combo.get()
        if data_file_path == "" or column == "" or sheet_name == "" or save_file == "":
            raise Exception("ERROR1:未入力項目があります。")
        if status == "追加":
            if not (os.path.exists(save_file + '/origin_data.npy') and os.path.exists(save_file + '/pca_data.npy') and os.path.exists(save_file + '/label_data.npy') and os.path.exists(save_file + '/sample.pickle')):
                raise Exception("ERROR14：セーブデータが見つからないため、データの追加割当を実行できません。\nクラスタの新規作成から行ってください。")
        column = int(column) - 1

class DataCleansing(ttk.Frame):
    """
    データクレンジングを行う
    """
    def __init__(self, root):
        super().__init__(root)
        self.delete_word_list = []
        self.mark_status = ""
        self.alpha_num_status = ""

    def create_frame(self):
        """
        ウィジェットの配置
        """
        global arr_data

        #トップフレーム
        self.top_frame = ttk.Frame(root)
        self.top_frame.pack(fill = tk.BOTH)

        #画面名ラベル
        title_label = ttk.Label(self.top_frame, text = "データクレンジング", font = ('', 20))
        title_label.pack(pady = (30, 0))
        subtitle_label = ttk.Label(self.top_frame, text = "データクレンジングを行います。")
        subtitle_label.pack(pady = 5)

        #treeview用フレーム
        tree_frame = ttk.Frame(self.top_frame)
        tree_frame.pack(fill = tk.BOTH, pady = (0, 10))
        #左カラムフレーム
        left_frame = ttk.Frame(self.top_frame)
        left_frame.pack(fill = tk.BOTH, side = tk.LEFT)
        #右カラムフレーム
        right_frame = ttk.Frame(self.top_frame)
        right_frame.pack(fill = tk.BOTH)
        #次へ、戻るボタン表示用フレーム
        button_frame = ttk.Frame(self.top_frame)
        button_frame.pack(anchor = tk.SE, padx = 30, pady = (120, 0))
        #エラーメッセージ表示用フレーム
        self.err_frame = tk.Frame(self.top_frame)
        self.err_frame.pack(pady = 10)

        #入力フォーム用フレーム
        entry_frame = ttk.Frame(left_frame)
        entry_frame.pack(padx = 20, pady = 10, anchor = tk.W)
        #記号、数字削除ボタン用フレーム
        delete_button_frame = ttk.Frame(left_frame)
        delete_button_frame.pack(padx = 20, pady = 10, anchor = tk.W)
        #フィルタリング情報のエクスポート用フレーム
        export_frame = ttk.Frame(right_frame)
        export_frame.pack(pady = 10, anchor = tk.N)

        #treeviewにスクロールバーを設置
        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side = tk.RIGHT, fill = tk.Y)
        self.tree = ttk.Treeview(tree_frame, yscrollcommand = scrollbar.set)
        scrollbar.config(command = self.tree.yview)

        self.tree["column"] = (1)
        self.tree["show"] = "headings"
        self.tree.heading(1, text = "セル値")

        insert_data = '"{}"'
        for cell_value in arr_data.tolist():
            cell_value = re.sub(r'["]', '”', cell_value)
            self.tree.insert("", "end", values = insert_data.format(cell_value))

        #treeview表示
        self.tree.pack(fill = tk.BOTH, expand = 1)

        #更新ボタン
        update_button = ttk.Button(tree_frame, text = "更新", command = self.update_tree)
        update_button.pack(anchor = tk.W)

        #デリートワード入力
        label1 = ttk.Label(entry_frame, text = "削除したいワードを入力してください")
        label1.grid(row = 0, column = 0, sticky = tk.W)
        self.txt_delete_word = ttk.Entry(entry_frame, width = 40)
        self.txt_delete_word.grid(row = 1, column = 0, sticky = tk.W)

        delete_button = ttk.Button(entry_frame, text = "削除", command = self.add_delete_word)
        delete_button.grid(row = 2, column = 0, sticky=tk.W)

        #削除したワードの表示
        label2 = ttk.Label(entry_frame, text = "削除したワードは以下です")
        label2.grid(row = 3, column = 0, sticky = tk.W)
        self.canvas = tk.Canvas(entry_frame, height = 70, bg="white")
        self.canvas.grid(row = 4, column = 0, sticky = tk.W)

        bar_y = ttk.Scrollbar(entry_frame, command = self.canvas.yview)
        bar_y.grid(row = 4, column = 1, sticky = tk.N + tk.S + tk.W)
        self.canvas.configure(yscrollcommand = bar_y.set)

        delete_word_frame = tk.Frame(self.canvas)
        self.delete_word_label = ttk.Label(delete_word_frame, text = 'まだありません', padding = (5, 5), background="white", wraplength = 250)
        self.delete_word_label.pack(anchor = tk.NW, fill = tk.BOTH)
        self.canvas.create_window((0,0), window = delete_word_frame, anchor = tk.NW)
        delete_word_frame.update()
        
        self.canvas.configure(scrollregion = (0, 0, self.delete_word_label.winfo_height(), self.delete_word_label.winfo_height()))

        #数字、記号削除ボタン
        delete_mark_button = ttk.Button(delete_button_frame, text = "記号削除", command = self.delete_mark)
        delete_mark_button.grid(row = 0, column = 0, sticky=tk.W)
        self.mark_label = tk.Label(delete_button_frame, text = self.mark_status)
        self.mark_label.grid(row = 1, column = 0, sticky=tk.W)

        delete_num_button = ttk.Button(delete_button_frame, text = "数字削除", command = self.delete_num)
        delete_num_button.grid(row = 2, column = 0, sticky=tk.W)
        self.num_label = tk.Label(delete_button_frame, text = self.alpha_num_status)
        self.num_label.grid(row = 3, column = 0, sticky=tk.W)

        self.show_filter()

        #csvへのエクスポート
        save_label = ttk.Label(export_frame, text = "削除したワードをcsvファイルに保存します。")
        save_label.grid(row = 0, column = 0, sticky=tk.W)
        save_button = ttk.Button(export_frame, text = "保存", command = self.save_delete_word)
        save_button.grid(row = 1, column = 0, sticky=tk.W)

        read_label = ttk.Label(export_frame, text = "csvファイルから削除するワードをインポートします。")
        read_label.grid(row = 2, column = 0, sticky=tk.W)
        self.txt_file_path = ttk.Entry(export_frame, width = 50)
        self.txt_file_path.grid(row = 3, column = 0, sticky = tk.W)
        button_selectfile = ttk.Button(export_frame, text = "参照", command = self.select_file)
        button_selectfile.grid(row = 3, column = 1, sticky = tk.W)
        read_button = ttk.Button(export_frame, text = "インポート", command = self.import_delete_word)
        read_button.grid(row = 4, column = 0, sticky=tk.W)

        #次へ、戻るボタン
        next_button = ttk.Button(button_frame, text = "次へ", command = self.next_page)
        next_button.grid(row = 0, column = 0)
        back_button = ttk.Button(button_frame, text = "戻る", command = self.back_page)
        back_button.grid(row = 0, column = 1)

        #エラーメッセージ用のラベル
        self.error_label = tk.Label(self.err_frame)
        self.error_label.pack()

    def add_delete_word(self):
        """
        デリートワードの追加
        """
        self.err_message = []

        delete_word = self.txt_delete_word.get()
        if delete_word == "":
            self.err_message.append("ERROR1:未入力項目があります。")
        else:
            self.delete_word_list.append(delete_word)

        if self.err_message:
            self.error_label.configure(text = "\n".join(self.err_message), fg = "red")

        self.show_filter()

    def show_filter(self):
        """
        デリートワードの表示
        """
        self.delete_word_label.configure(text = self.delete_word_list)
        self.canvas.configure(scrollregion = (0, 0, self.delete_word_label.winfo_height(), self.delete_word_label.winfo_height()))

        self.mark_label.configure(text = self.mark_status)
        self.num_label.configure(text = self.alpha_num_status)

    def delete_word(self):
        """
        デリートワードの削除
        """
        global arr_data
        data_list = []
        for data in arr_data.tolist():
            for delete_word in self.delete_word_list:
                data = re.sub(delete_word, '', data)
            data_list.append(data)
        arr_data = np.array(data_list)

    def delete_mark(self):
        """
        記号の削除
        """
        global arr_data
        data_list = []
        for data in arr_data.tolist():
            data = re.sub(r'[!"#$%&()\*\+\-\.,\/:;<=>?@\[\\\]^_`{|}~]', ' ', data)
            data = re.sub(r'[︰-＠“”≪≫〈〉《》『』※、。⇒・★☆■【】①②③「」→]', ' ', data)
            data_list.append(data)
        arr_data = np.array(data_list)

        self.mark_status = "記号を削除しました。"

        self.show_filter()

    def delete_num(self):
        """
        数字の削除
        """
        global arr_data
        data_list = []
        for data in arr_data.tolist():
            data = re.sub(r'[0-9]', '', data)
            data_list.append(data)
        arr_data = np.array(data_list)

        self.alpha_num_status = "数字を削除しました。"

        self.show_filter()

    def save_delete_word(self):
        """
        保存ボタン押下時の処理
        """
        self.err_message = []

        iDir = os.path.abspath(os.path.dirname(__file__))
        file_path = tk.filedialog.asksaveasfilename(initialdir = iDir, filetypes = [("CSV Files", "*.csv")], defaultextension = 'csv')

        if file_path:
            try:
                with open(file_path, 'w') as f:
                    writer = csv.writer(f)
                    writer.writerow(self.delete_word_list)
            except PermissionError as e:
                self.err_message.append("ERROR5:csvへの出力時にエラーが発生しました。csvファイルを閉じているか確認してください。")
            else:
                tkm.showinfo("", "csvの出力が完了しました。")

        if self.err_message:
            self.error_label.configure(text = "\n".join(self.err_message), fg = "red")

    def import_delete_word(self):
        """
        インポートボタン押下時の処理
        """
        self.err_message = []

        try:
            file_path = self.txt_file_path.get()
            with open(file_path) as f:
                reader = csv.reader(f)
                content = [row for row in reader]
            self.delete_word_list[len(self.delete_word_list):len(self.delete_word_list)] = content[0]
            self.delete_word_list = list(set(self.delete_word_list))
        except FileNotFoundError as e:
            self.err_message.append("ERROR3:ご指定のファイルが見つかりません。ファイルパスを確認してください。")
        else:
            self.update_tree()

        if self.err_message:
            self.error_label.configure(text = "\n".join(self.err_message), fg = "red")

    def select_file(self):
        """
        参照ボタン押下時の処理
        """
        iDir = os.path.abspath(os.path.dirname(__file__))
        file_path = tk.filedialog.askopenfilename(initialdir = iDir, filetypes = [("CSV Files", "*.csv")])
        self.txt_file_path.insert(tk.END, file_path)

    def update_tree(self):
        """
        treeviewを更新する
        """
        self.canvas.delete("all")
        self.delete_word()
        self.top_frame.destroy()
        self.create_frame()

    def next_page(self):
        """
        次へボタン押下時の処理
        """
        self.top_frame.destroy()

        ud = UserDict(root)
        ud.create_frame()

    def back_page(self):
        """
        戻るボタン押下時の処理
        """
        self.top_frame.destroy()

        ef = EntryForm(root)
        ef.create_widget()

class UserDict(ttk.Frame):
    """
    ユーザー辞書の作成を行う
    """
    def __init__(self, root):
        super().__init__(root)

    def create_frame(self):
        """
        ウィジェットの配置
        """
        self.top_frame = ttk.Frame(root)
        self.top_frame.pack(fill = tk.BOTH)

        #画面名ラベル
        title_label = ttk.Label(self.top_frame, text = "ユーザー辞書作成", font = ('', 20))
        title_label.pack(padx = 5, pady = (30, 0))
        subtitle_label = ttk.Label(self.top_frame, text = "ユーザー辞書の自動生成を行います。")
        subtitle_label.pack(padx = 5, pady = 5)

        createdic_frame = ttk.Frame(self.top_frame)
        createdic_frame.pack(pady = (200, 0))

        file_frame = ttk.Frame(self.top_frame)
        file_frame.pack(pady = 20)

        execute_frame = ttk.Frame(self.top_frame)
        execute_frame.pack(pady = 20)

        button_frame = ttk.Frame(self.top_frame)
        button_frame.pack(anchor = tk.SE, padx = 30, pady = (110, 0))

        self.err_frame = tk.Frame(self.top_frame)
        self.err_frame.pack()

        label = ttk.Label(createdic_frame, text = "実行ボタンを押すとユーザー辞書の自動生成が開始します。")
        label.pack()
        #ユーザー辞書作成ボタンの表示
        execute_button = ttk.Button(createdic_frame, text = "実行", command = self.create_user_dict)
        execute_button.pack()

        label2 = ttk.Label(file_frame, text = "ユーザー辞書(CSV)")
        label2.grid(row = 0, column = 0, sticky=tk.W)

        self.txt_file_path = ttk.Entry(file_frame, width = 50)
        self.txt_file_path.grid(row = 1, column = 0, columnspan = 2)
        button_selectfile = ttk.Button(file_frame, text = "参照", command = self.select_file)
        button_selectfile.grid(row = 1, column = 2)
        label3 = ttk.Label(file_frame, text = "ユーザー辞書を使用しない場合は空欄にしてください。")
        label3.grid(row = 2, column = 0, sticky=tk.W, columnspan = 2)


        next_button = ttk.Button(execute_frame, text = "形態素解析の実行", command = self.next_page)
        next_button.pack()

        #戻るボタン表示
        back_button = ttk.Button(button_frame, text = "戻る", command = self.back_page)
        back_button.pack(side = "left")

        self.error_label = tk.Label(self.err_frame)
        self.error_label.pack()

    def select_file(self):
        """
        参照ボタン押下時の処理
        """
        iDir = os.path.abspath(os.path.dirname(__file__))
        file_path = tk.filedialog.askopenfilename(initialdir = iDir, filetypes = [("CSV Files", "*.csv")])
        self.txt_file_path.insert(tk.END, file_path)

    def create_user_dict(self):
        """
        ユーザー辞書の作成を行う
        """
        self.err_message = []
        self.t = Tokenizer()
        global arr_data

        corpus_phrase = []
        for data in arr_data.tolist():
            corpus = []
            #形態素解析を実行
            tokens = self.t.tokenize(data)
            for token in tokens:
                if token.surface != "":
                    corpus.append(token.surface)
            corpus_phrase.append(corpus)

        #複合語候補を作成
        for i in range(2):
            phrases = Phrases(corpus_phrase, min_count = 2, threshold = 4.0)
            bigram = Phraser(phrases)
            transformed_bi = list(bigram[corpus_phrase])
            corpus_phrase = transformed_bi

        del phrases
        del bigram
        gc.collect()

        words_df_noun = pd.DataFrame()
        #複合語の判定
        for sentence in transformed_bi:
            for word in sentence:
                boolean_noun = self.judge_noun(word)
                word = word.replace("_","")
                if boolean_noun:
                    words_df_noun = words_df_noun.append([word])

        if not words_df_noun.empty:
            #ユーザー辞書のフォーマット通りにcsvを作っていく
            words_df_noun.columns=["複合語"]
            words_df_noun = words_df_noun.groupby("複合語").sum().reset_index()
            words_df_noun["paramater"]="-1,-1,1000,名詞,一般,*,*,*,*,%s,*,*"
            words_df_noun = pd.concat([words_df_noun['複合語'],words_df_noun['paramater'].str.split(',', expand=True)], axis=1)
            words_df_noun[9] = words_df_noun["複合語"]

            try:
                iDir = os.path.abspath(os.path.dirname(__file__))
                file_path = tk.filedialog.asksaveasfilename(initialdir = iDir, filetypes = [("CSV Files", "*.csv"), ("すべて", "*")])
                words_df_noun.to_csv(file_path, sep=",", index=False, header=False, encoding='cp932')
                self.txt_file_path.insert(tk.END, file_path)

            except PermissionError as e:
                self.err_message.append("ERROR5:csvへの出力時にエラーが発生しました。csvファイルを閉じているか確認してください。")
            else:
                tkm.showinfo("", "ユーザー辞書の作成が完了しました")
        
            if self.err_message:
                self.error_label.configure(text = "\n".join(self.err_message), fg = "red")
        else:
            tkm.showinfo("", "複合語が見つからなかったため、ユーザー辞書を作成しませんでした。")

        del words_df_noun
        del self.t
        gc.collect()

    def judge_noun(self, word):
        """
        複合名詞を抽出

        Parameter
        ---------
        word：str
            複合語の候補
        """
        if word.find('_') > 0:
            word = word.replace("_","")
            if re.search(r'[ 　]', word):
                return False
            if re.match(r'^[あ-ん]+$', word) and len(word) <= 2:
                return False
            if re.match(r'[a-z]', word):
                return False
            tokens = self.t.tokenize(word)
        else:
            return False
        for num, token in enumerate(tokens):
            pos_part0 = token.part_of_speech.split(',')[0]
            pos_part1 = token.part_of_speech.split(',')[1]
            pos_part2 = token.part_of_speech.split(',')[2]
            #名詞以外はNG
            if pos_part0 != "名詞":
                return False
            #数字、助数字が含まれていたらNG
            if pos_part0 == "名詞" and (pos_part1 == u'数' or pos_part2 == "助数詞"):
                return False
        return True

    def next_page(self):
        """
        次へボタン押下時の処理
        """
        global userdict_file_path

        self.err_message = []
        userdict_file_path = self.txt_file_path.get() #読み込みファイルのパス

        #形態素解析
        try:
            clustering = Clustering()
            clustering.morphological()
        except FileNotFoundError as e:
            self.err_message.append("ERROR6:ユーザー辞書が見つかりません。")
        except ValueError as e:
            self.err_message.append("ERROR7:ユーザー辞書に誤りがあります。")
        else:
            self.top_frame.destroy()
            mv = MorphologicalVisualization(root)
            mv.show_tree()

        if self.err_message:
            self.error_label.configure(text = "\n".join(self.err_message), fg = "red")

    def back_page(self):
        """
        戻るボタン押下時の処理
        """
        self.top_frame.destroy()
        df = DataCleansing(root)
        df.create_frame()

class MorphologicalVisualization(ttk.Frame):
    """
    形態素解析の結果を表示し、品詞やストップワードでフィルタリングを行う
    """
    def __init__(self, root):
        super().__init__(root)

        self.clicked_column = ""
        self.text_list = []
        self.stopwords = []
        self.pos_list = []
        self.unify_dict = {}

    def show_tree(self):
        """
        tree_viewを表示する
        """
        global all_token

        #トップフレーム
        self.top_frame = ttk.Frame(root)
        self.top_frame.pack(fill = tk.BOTH)

        #画面名ラベル
        title_label = ttk.Label(self.top_frame, text = "形態素解析結果のフィルタリング", font = ('', 20))
        title_label.pack(padx = 5, pady = (30, 0))
        subtitle_label = ttk.Label(self.top_frame, text = "形態素解析結果から、ストップワードの選択や品詞のフィルタリングを行います。")
        subtitle_label.pack(padx = 5, pady = 5)
        
        #treeview表示用のフレーム
        tree_frame = ttk.Frame(self.top_frame)
        tree_frame.pack(fill = tk.BOTH)
        #更新ボタン表示用のフレーム
        update_frame = ttk.Frame(self.top_frame)
        update_frame.pack(anchor = tk.W, padx = 20, pady = 10)
        #フィルタリング関係をいれるフレーム
        filter_frame = ttk.Frame(self.top_frame)
        filter_frame.pack(anchor = tk.W, padx = 20, pady = 10)
        #フィルター表示用のフレーム
        display_frame = ttk.Frame(filter_frame)
        display_frame.pack(side = tk.LEFT)
        #フィルタリング情報のエクスポート
        export_frame = ttk.Frame(filter_frame)
        export_frame.pack(side = tk.LEFT, anchor = tk.N, padx = 20)
        #次へ、戻るボタン表示用のフレーム
        button_frame = ttk.Frame(self.top_frame)
        button_frame.pack(anchor = tk.SE, padx = 30)
        #エラーメッセージ表示用のフレーム
        self.err_frame = tk.Frame(self.top_frame)
        self.err_frame.pack(pady = 10)

        #treeviewにスクロールバーを設置
        scrollbar = ttk.Scrollbar(tree_frame)
        scrollbar.pack(side = tk.RIGHT, fill = tk.Y)
        self.tree = ttk.Treeview(tree_frame, yscrollcommand = scrollbar.set)
        scrollbar.config(command = self.tree.yview)

        #カラム
        self.tree["column"] = (1, 2, 3)
        self.tree["show"] = "headings"
        self.tree.heading(1, text = "単語", command = self.column_click)
        self.tree.heading(2, text = "品詞", command = self.column_click)
        self.tree.heading(3, text = "出現回数", command = self.column_click)

        #単語の出現頻度を数える
        words = []
        for token_dict in all_token:
            for key in token_dict.keys():
                words.append(key)
        count = collections.Counter(words)

        #データの挿入(重複している単語は省いて表示)
        display_dict = {}
        for token_dict in all_token:
            display_dict.update(token_dict)
        for key in display_dict.keys():
            display_dict[key] = [display_dict[key], count[key]]

        #ソート機能
        if self.clicked_column != "":
            if self.clicked_column == "#1":
                #単語でソート
                sort = sorted(display_dict.items(), key = lambda x : x[0])
            elif self.clicked_column == "#2":
                #品詞でソート
                sort = sorted(display_dict.items(), key = lambda x : x[1][0])
            elif self.clicked_column == "#3":
                #出現頻度でソート
                sort = sorted(display_dict.items(), key = lambda x : x[1][1])
            for sort_item in sort:
                self.tree.insert("", "end", values = (sort_item[0], sort_item[1][0], sort_item[1][1]))
        else:
            for key in display_dict.keys():
                self.tree.insert("", "end", values = (key, display_dict[key][0], display_dict[key][1]))

        #treeview表示
        self.tree.pack(fill = tk.BOTH, expand = 1)

        #右クリック時のイベント
        self.tree.bind("<Button-3>", self.selected)

        #更新ボタン設置
        update_button = ttk.Button(update_frame, text = "更新", command = self.update_tree)
        update_button.grid(row = 0, column = 0)

        #ストップワード表示
        stopword_label = ttk.Label(display_frame, text = "追加したストップワードは以下です。")
        stopword_label.grid(row = 0, column = 0, sticky=tk.W)
        self.canvas2 = tk.Canvas(display_frame, height = 70, bg="white")
        self.canvas2.grid(row = 1, column = 0, sticky = tk.W)

        bar_y2 = ttk.Scrollbar(display_frame, command = self.canvas2.yview)
        bar_y2.grid(row = 1, column = 1, sticky = tk.N + tk.S + tk.W)
        self.canvas2.configure(yscrollcommand = bar_y2.set)

        stopword_frame = tk.Frame(self.canvas2)
        self.stopword_item_label = ttk.Label(stopword_frame, text = "まだありません", padding = (5, 5), background="white", wraplength = 250)
        self.stopword_item_label.pack(anchor = tk.NW, fill = tk.BOTH)
        self.display_stopword()
        self.canvas2.create_window((0,0), window = stopword_frame, anchor = tk.NW)
        stopword_frame.update()

        self.canvas2.configure(scrollregion = (0, 0, self.stopword_item_label.winfo_height(), self.stopword_item_label.winfo_height()))

        #品詞表示
        pos_label = ttk.Label(display_frame, text = "排除した品詞は以下です。")
        pos_label.grid(row = 2, column = 0, sticky = tk.W)
        self.canvas3 = tk.Canvas(display_frame, height = 70, bg="white")
        self.canvas3.grid(row = 3, column = 0, sticky = tk.W)

        bar_y3 = ttk.Scrollbar(display_frame, command = self.canvas3.yview)
        bar_y3.grid(row = 3, column = 1, sticky = tk.N + tk.S + tk.W)
        self.canvas3.configure(yscrollcommand = bar_y3.set)

        pos_frame = tk.Frame(self.canvas3)
        self.pos_item_label = ttk.Label(pos_frame, text = "まだありません", padding = (5, 5), background="white", wraplength = 250)
        self.pos_item_label.pack(anchor = tk.NW, fill = tk.BOTH)
        self.display_pos()
        self.canvas3.create_window((0,0), window = pos_frame, anchor = tk.NW)
        pos_frame.update()

        self.canvas3.configure(scrollregion = (0, 0, self.pos_item_label.winfo_height(), self.pos_item_label.winfo_height()))

        #表記ゆれの統一をした単語の表示
        unify_label = ttk.Label(display_frame, text = "表記ゆれの統一を行った単語は以下です。")
        unify_label.grid(row = 4, column = 0, sticky = tk.W)
        self.canvas = tk.Canvas(display_frame, height = 70, bg="white")
        self.canvas.grid(row = 5, column = 0, sticky = tk.W)

        bar_y = ttk.Scrollbar(display_frame, command = self.canvas.yview)
        bar_y.grid(row = 5, column = 1, sticky = tk.N + tk.S + tk.W)
        self.canvas.configure(yscrollcommand = bar_y.set)

        unify_frame = tk.Frame(self.canvas)
        self.unify_label = ttk.Label(unify_frame, text = 'まだありません', padding = (5, 5), background="white", wraplength = 250)
        self.unify_label.pack(anchor = tk.NW, fill = tk.BOTH)
        self.display_unify()
        self.canvas.create_window((0,0), window = unify_frame, anchor = tk.NW)
        unify_frame.update()

        self.canvas.configure(scrollregion = (0, 0, self.unify_label.winfo_height(), self.unify_label.winfo_height()))

        #ストップワードと品詞の保存、読込エリア
        save_label = ttk.Label(export_frame, text = "行ったデータ整形をcsvファイルに保存します。")
        save_label.grid(row = 6, column = 0, sticky=tk.W)
        save_button = ttk.Button(export_frame, text = "保存", command = self.save_filter)
        save_button.grid(row = 7, column = 0, sticky=tk.W)
        read_label = ttk.Label(export_frame, text = "保存済のデータ整形をインポートします。")
        read_label.grid(row = 8, column = 0, sticky=tk.W)
        self.txt_file_path = ttk.Entry(export_frame, width = 50)
        self.txt_file_path.grid(row = 9, column = 0, sticky = tk.W)
        button_selectfile = ttk.Button(export_frame, text = "参照", command = self.select_file)
        button_selectfile.grid(row = 9, column = 1, sticky = tk.W)
        read_button = ttk.Button(export_frame, text = "インポート", command = self.import_filter)
        read_button.grid(row = 10, column = 0, sticky=tk.W)

        #実行ボタン表示
        next_button = ttk.Button(button_frame, text = "次へ", command = self.next_page)
        next_button.grid(row = 0, column = 2)

        back_buttom = ttk.Button(button_frame, text = "戻る", command = self.back_page)
        back_buttom.grid(row = 0, column = 3)

        #エラーメッセージ用のラベル
        self.error_label = tk.Label(self.err_frame)
        self.error_label.pack()

    def column_click(self):
        """
        クリックされたカラムを取得する
        """
        x = self.tree.winfo_pointerx() - self.tree.winfo_rootx()
        self.clicked_column = self.tree.identify_column(x)
        self.update_tree()

    def save_filter(self):
        """
        保存ボタン押下時の処理
        """
        self.err_message = []

        iDir = os.path.abspath(os.path.dirname(__file__))
        file_path = tk.filedialog.asksaveasfilename(initialdir = iDir, filetypes = [("CSV Files", "*.csv")], defaultextension = 'csv')

        if file_path:
            try:
                with open(file_path, 'w') as f:
                    writer = csv.writer(f)
                    writer.writerow(self.stopwords)
                    writer.writerow(self.pos_list)
                    writer.writerow(list(self.unify_dict.keys()))
                    writer.writerow(list(self.unify_dict.values()))
            except PermissionError as e:
                self.err_message.append("ERROR5:csvへの出力時にエラーが発生しました。csvファイルを閉じているか確認してください。")
            else:
                tkm.showinfo("", "csvの出力が完了しました。")

        if self.err_message:
            self.error_label.configure(text = "\n".join(self.err_message), fg = "red")

    def import_filter(self):
        """
        インポートボタン押下時の処理
        """
        self.err_message = []

        try:
            file_path = self.txt_file_path.get()
            with open(file_path) as f:
                reader = csv.reader(f)
                content = [row for row in reader]
            self.stopwords[len(self.stopwords):len(self.stopwords)] = content[0]
            self.stopwords = list(set(self.stopwords))
            self.pos_list[len(self.pos_list):len(self.pos_list)] = content[2]
            self.pos_list = list(set(self.pos_list))
            for key, word in zip(content[4], content[6]):
                word = re.sub(r'["\'[\\\]]', '', word)
                word = re.sub(r'[ 　]', '', word)
                word = word.split(",")
                self.unify_dict[key] = word
        except FileNotFoundError as e:
            self.err_message.append("ERROR3:ご指定のファイルが見つかりません。ファイルパスを確認してください。")
        else:
            self.update_tree()

        if self.err_message:
            self.error_label.configure(text = "\n".join(self.err_message), fg = "red")

    def select_file(self):
        """
        参照ボタン押下時の処理
        """
        iDir = os.path.abspath(os.path.dirname(__file__))
        file_path = tk.filedialog.askopenfilename(initialdir = iDir, filetypes = [("CSV Files", "*.csv")])
        self.txt_file_path.insert(tk.END, file_path)

    def selected(self, event):
        """
        ポップアップメニューを表示する
        """
        #選択中の行を取得
        curItems = self.tree.selection()
        self.selected_items = [self.tree.item(i)['values'] for i in curItems]
        
        popup_menu = tk.Menu(self.tree, tearoff = 0)
        popup_menu.add_command(label = "ストップワードを追加", command = self.entry_stopword)
        popup_menu.add_command(label = "品詞でフィルタ", command = self.filter_pos)
        popup_menu.add_command(label = "表記ゆれの統一", command = self.entry_unify)
        popup_menu.post(event.x_root,event.y_root)

    def entry_unify(self):
        """
        表記ゆれを統一する単語の入力
        """
        self.unify_window = tk.Tk()
        self.unify_window.title("表記ゆれの統一")
        self.unify_window.geometry("300x200")

        word_frame = tk.Frame(self.unify_window)
        word_frame.pack(padx = 10, pady = 10)
        entry_frame = tk.Frame(self.unify_window)
        entry_frame.pack(padx = 10, pady = 10)
        button_frame = tk.Frame(self.unify_window)
        button_frame.pack(anchor = tk.SE, padx = 10, pady = 20)

        for item in self.selected_items:
            word_label = ttk.Label(word_frame, text = item[0],  borderwidth = 1, relief = "solid", background = "#fff", padding = (10, 5))
            word_label.pack(side = "left", padx = 5)

        msg_label = ttk.Label(entry_frame, text = "上記の単語を指定の表記に統一します")
        msg_label.pack()
        self.txt_unify_word = ttk.Entry(entry_frame, width = 30)
        self.txt_unify_word.pack()

        ok_button = ttk.Button(button_frame, text = "実行", command = self.add_unify_dict)
        ok_button.pack(side = "left")
        cancel_button = ttk.Button(button_frame, text = "キャンセル", command = partial(self.close_window, self.unify_window))
        cancel_button.pack(side = "left")

    def add_unify_dict(self):
        unify_word = self.txt_unify_word.get()
        select_words = []

        for item in self.selected_items:
            select_words.append(item[0])

        self.unify_dict[unify_word] = list(set(select_words))

        self.close_window(self.unify_window)
        self.display_unify()

    def execute_unify(self):
        """
        表記ゆれの統一の実行
        """
        global all_token

        for token_dict in all_token:
            for key in list(token_dict):
                for unify_word in list(self.unify_dict):
                    if key in self.unify_dict[unify_word]:
                        token_dict[unify_word] = token_dict.pop(key)

    def display_unify(self):
        """
        表記ゆれの統一を行った単語を表示する
        """
        for key in self.unify_dict.keys():
            text = "「" + '、'.join(self.unify_dict[key]) + "」を「" + key + "」に統一しました。"
            self.text_list.append(text)

        self.text_list = list(set(self.text_list))

        if self.text_list:
            self.unify_label.configure(text = "\n".join(self.text_list))
            self.canvas.configure(scrollregion = (0, 0, self.unify_label.winfo_height(), self.unify_label.winfo_height()))

    def entry_stopword(self):
        """
        ストップワードを追加するか確認するウィンドウ
        """
        self.stopword_window = tk.Tk()
        self.stopword_window.title("ストップワードの追加")
        self.stopword_window.geometry("300x200")

        word_frame = tk.Frame(self.stopword_window)
        word_frame.pack(padx = 10, pady = 10)
        sub_frame = tk.Frame(self.stopword_window)
        sub_frame.pack()
        button_frame = tk.Frame(self.stopword_window)
        button_frame.pack(anchor = tk.SE, padx = 10, pady = 20)

        for item in self.selected_items:
            word_label = ttk.Label(word_frame, text = item[0],  borderwidth = 1, relief = "solid", background = "#fff", padding = (10, 5))
            word_label.pack(side = "left", padx = 5)
        msg_label = ttk.Label(sub_frame, text = "この単語をストップワードに追加しますか？")
        msg_label.pack()
        ok_button = ttk.Button(button_frame, text = "追加", command = self.add_stopword)
        ok_button.pack(side = "left")
        cancel_button = ttk.Button(button_frame, text = "キャンセル", command = partial(self.close_window, self.stopword_window))
        cancel_button.pack(side = "left")

    def add_stopword(self):
        """
        ストップワードをリストに追加する
        """
        for item in self.selected_items:
            self.stopwords.append(item[0])
        self.stopwords = list(set(self.stopwords))
        self.close_window(self.stopword_window)
        self.display_stopword()

    def display_stopword(self):
        """
        追加したストップワードを表示する
        """
        if self.stopwords:
            self.stopword_item_label.configure(text = self.stopwords)
            self.canvas2.configure(scrollregion = (0, 0, self.stopword_item_label.winfo_height(), self.stopword_item_label.winfo_height()))

    def remove_stopword(self):
        """
        ストップワードを表から排除する
        """
        global all_token
        for token_dict in all_token:
            for stopword in self.stopwords:
                if stopword in token_dict:
                    del token_dict[stopword]

    def filter_pos(self):
        """
        品詞を排除するか確認するウィンドウ
        """
        self.pos_window = tk.Tk()
        self.pos_window.title("品詞のフィルタリング")
        self.pos_window.geometry("300x200")

        word_frame = tk.Frame(self.pos_window)
        word_frame.pack(padx = 10, pady = 10)
        sub_frame = tk.Frame(self.pos_window)
        sub_frame.pack()
        button_frame = tk.Frame(self.pos_window)
        button_frame.pack(anchor = tk.SE, padx = 10, pady = 20)

        for item in self.selected_items:
            word_label = ttk.Label(word_frame, text = item[1],  borderwidth = 1, relief = "solid", background = "#fff", padding = (10, 5))
            word_label.pack(side = "left",padx = 5)
        msg_label = ttk.Label(sub_frame, text = "この品詞を排除しますか？")
        msg_label.pack()
        ok_button = ttk.Button(button_frame, text = "排除", command = self.add_pos)
        ok_button.pack(side = "left")
        cancel_button = ttk.Button(button_frame, text = "キャンセル", command = partial(self.close_window, self.pos_window))
        cancel_button.pack(side = "left")

    def add_pos(self):
        """
        排除する品詞のリストを作成
        """
        for item in self.selected_items:
            self.pos_list.append(item[1])
        self.pos_list = list(set(self.pos_list))
        self.close_window(self.pos_window)
        self.display_pos()

    def display_pos(self):
        """
        排除する品詞のリストを表示
        """
        if self.pos_list:
            self.pos_item_label.configure(text = self.pos_list)
            self.canvas3.configure(scrollregion = (0, 0, self.pos_item_label.winfo_height(), self.pos_item_label.winfo_height()))

    def remove_pos(self):
        """
        選択された品詞を排除する
        """
        global all_token
        for token_dict in all_token:
            for pos in self.pos_list:
                keys = [k for k, v in token_dict.items() if v == pos]
                for key in keys:
                    del token_dict[key]

    def close_window(self, closewindow):
        """
        ウィンドウを閉じる
        """
        closewindow.destroy()

    def back_page(self):
        """
        戻るボタン押下時の処理
        """
        self.top_frame.destroy()
        ud = UserDict(root)
        ud.create_frame()

    def next_page(self):
        """
        次へボタン押下時の処理
        """
        global status

        self.segmentation()
        self.top_frame.destroy()

        if status == "新規":
            sk = SelectKvalue(root)
            sk.entry_kvalue()
        elif status == "追加":
            aec = AssignExistingCluster(root)
            aec.create_widget()

    def segmentation(self):
        """
        フィルタリング後のデータを分かち書きにする
        """
        global all_token
        global segmentation_datas
        all_token_surfaces = []
        segmentation_datas = []
        for token_dict in all_token:
            token_surfaces = list(token_dict.keys())
            segmentation_data = ' '.join(token_surfaces)
            segmentation_datas.append(segmentation_data)

    def delete_tree(self):
        """
        treeviewを削除する
        """
        self.canvas.delete("all")
        self.canvas2.delete("all")
        self.canvas3.delete("all")
        self.top_frame.destroy()

    def update_tree(self):
        """
        treeviewを更新する
        """
        self.remove_stopword()
        self.remove_pos()
        self.execute_unify()
        self.delete_tree()
        self.show_tree()

class AssignExistingCluster(ttk.Frame):
    """
    既存のクラスタにデータを割り当てる
    """
    def __init__(self, root):
        super().__init__(root)

    def create_widget(self):
        """
        ウィジェットの配置
        """
        self.top_frame = ttk.Frame(root)
        self.top_frame.pack(fill = tk.BOTH)

        #画面名ラベル
        title_label = ttk.Label(self.top_frame, text = "クラスタリング", font = ('', 20))
        title_label.pack(padx = 5, pady = (30, 0))
        subtitle_label = ttk.Label(self.top_frame, text = "クラスタリングを実行します。")
        subtitle_label.pack(padx = 5, pady = 5)

        clustering_frame = ttk.LabelFrame(self.top_frame, text = "クラスタリング")
        clustering_frame.pack(fill = tk.BOTH, padx = 20, pady = 20, ipadx = 10, ipady = 10)

        button_frame = ttk.Frame(self.top_frame)
        button_frame.pack(anchor = tk.SE, padx = 30, pady = 20)

        self.err_frame = tk.Frame(self.top_frame)
        self.err_frame.pack(pady = 10)

        button_execute = ttk.Button(clustering_frame, text = "クラスタリングを実行する", command = self.call_clustering_func)
        button_execute.grid(row = 2, column = 0)

        top_button = ttk.Button(button_frame, text = "トップへ", command = self.go_top_page)
        top_button.grid(row = 0, column = 0)
        back_buttom = ttk.Button(button_frame, text = "戻る", command = self.back_page)
        back_buttom.grid(row = 0, column = 1)

        self.error_label = tk.Label(self.err_frame)
        self.error_label.pack()

    def call_clustering_func(self):
        """
        クラスタリングの処理を呼び出す
        """
        self.err_message = []
        try:
            clustering = Clustering()
            pca_data = clustering.digitization()
            labels, sse, ranking_tops = clustering.assign_cluster(pca_data)
            clustering.output_excel(labels, sse, ranking_tops)
        except PermissionError as e:
            self.err_message.append("ERROR10:excelへの出力時にエラーが発生しました。excelファイルを閉じているか確認してください。")
        else:
            tkm.showinfo("", "クラスタリング結果をresult.xlsxに出力しました。")

        if self.err_message:
            self.error_label.configure(text = "\n".join(self.err_message), fg = "red")

    def go_top_page(self):
        """
        トップへボタン押下時の処理
        """
        self.top_frame.destroy()
        ef = EntryForm(root)
        ef.create_widget()

    def back_page(self):
        """
        戻るボタン押下時の処理
        """
        self.top_frame.destroy()
        mv = MorphologicalVisualization(root)
        mv.show_tree()

class SelectKvalue(ttk.Frame):
    """
    エルボー法とクラスタリングの実行を行う
    """
    def __init__(self, root):
        super().__init__(root)

    def entry_kvalue(self):
        """
        ウィジェットの配置
        """
        self.top_frame = ttk.Frame(root)
        self.top_frame.pack(fill = tk.BOTH)

        #画面名ラベル
        title_label = ttk.Label(self.top_frame, text = "クラスタリング", font = ('', 20))
        title_label.pack(padx = 5, pady = (30, 0))
        subtitle_label = ttk.Label(self.top_frame, text = "K値を決めて、クラスタリングを実行します。")
        subtitle_label.pack(padx = 5, pady = 5)
        
        elbow_frame = ttk.LabelFrame(self.top_frame, text = "エルボー法")
        elbow_frame.pack(fill = tk.BOTH, padx = 20, pady = 20, ipadx = 10, ipady = 10)

        clustering_frame = ttk.LabelFrame(self.top_frame, text = "クラスタリング")
        clustering_frame.pack(fill = tk.BOTH, padx = 20, pady = 20, ipadx = 10, ipady = 10)

        button_frame = ttk.Frame(self.top_frame)
        button_frame.pack(anchor = tk.SE, padx = 30, pady = (230, 0))

        self.err_frame = tk.Frame(self.top_frame)
        self.err_frame.pack(pady = 10)

        kvalue_label = ttk.Label(clustering_frame, text = "K値")
        kvalue_label.grid(column = 0, row = 0, sticky = tk.W)
        self.txt_kvalue = ttk.Entry(clustering_frame, width = 10)
        self.txt_kvalue.grid(column = 1, row = 0, sticky = tk.W)

        execute_button = ttk.Button(clustering_frame, text = "クラスタリングを実行する", command = self.call_clustering_func)
        execute_button.grid(column = 0, row = 1, columnspan = 3, sticky = tk.W)

        elbow_label = ttk.Label(elbow_frame, text = "エルボー法を実行して最適なK値を分析できます。")
        elbow_label.grid(column = 0, row = 0, columnspan = 5, sticky = tk.W)

        elbow_kvalue_label = ttk.Label(elbow_frame, text = "最大のK値")
        elbow_kvalue_label.grid(column = 0, row = 1, sticky = tk.W)
        self.txt_elbow_kvalue = ttk.Entry(elbow_frame, width = 10)
        self.txt_elbow_kvalue.grid(column = 1, row = 1, sticky = tk.W)

        elbow_button = ttk.Button(elbow_frame, text = "エルボー法を実行する", command = self.call_elbow_func)
        elbow_button.grid(column = 0, row = 2, columnspan = 3, sticky = tk.W)

        top_button = ttk.Button(button_frame, text = "トップへ", command = self.go_top_page)
        top_button.grid(row = 0, column = 0)
        back_button = ttk.Button(button_frame, text = "戻る", command = self.back_page)
        back_button.grid(row = 0, column = 1)


        self.error_label = tk.Label(self.err_frame)
        self.error_label.pack()

    def call_clustering_func(self):
        """
        クラスタリングの処理を呼び出す
        """
        self.err_message = []
        #入力されたK値の取得
        k_value = self.txt_kvalue.get()
        if k_value == "":
            self.err_message.append("ERROR8:K値が未入力です。")

        try:
            k_value = int(k_value)
            clustering = Clustering()
            pca_data = clustering.digitization()
            labels, sse, ranking_tops = clustering.clustering(pca_data, k_value)
            clustering.output_excel(labels, sse, ranking_tops)
        except ValueError as e:
            self.err_message.append("ERROR9:K値は数字で入力してください。")
        except PermissionError as e:
            self.err_message.append("ERROR10:excelへの出力時にエラーが発生しました。excelファイルを閉じているか確認してください。")
        else:
            tkm.showinfo("", "クラスタリング結果をresult.xlsxに出力しました。")

        if self.err_message:
            self.error_label.configure(text = "\n".join(self.err_message), fg = "red")

    def call_elbow_func(self):
        """
        エルボー法の処理を呼び出す
        """
        self.err_message = []
        #入力された最大のK値の取得
        elbow_kvalue = self.txt_elbow_kvalue.get()
        if elbow_kvalue == "":
            self.err_message.append("ERROR11:最大のK値が未入力です。")
        try:
            elbow_kvalue = int(elbow_kvalue) + 1
        except ValueError as e:
            self.err_message.append("ERROR12:最大のK値は数字で入力してください。")
        else:
            clustering = Clustering()
            pca_data = clustering.digitization()
            clustering.elbow(pca_data, elbow_kvalue)

        if self.err_message:
            self.error_label.configure(text = "\n".join(self.err_message), fg = "red")

    def go_top_page(self):
        """
        トップへボタン押下時の処理
        """
        self.top_frame.destroy()
        ef = EntryForm(root)
        ef.create_widget()

    def back_page(self):
        """
        戻るボタン押下時の処理
        """
        self.top_frame.destroy()
        mv = MorphologicalVisualization(root)
        mv.show_tree()

if __name__ == '__main__':
    """
    ウィンドウの表示とウィジェットの配置
    """
    root = tk.Tk()
    root.title("クラスタリング")
    root.geometry("860x700")
    ef = EntryForm(root)
    ef.create_widget()
    root.mainloop()
